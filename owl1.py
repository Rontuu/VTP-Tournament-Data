import requests
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sortedcontainers import SortedList
import datetime
import time
import arrow

#reference:
#matchhistory
#https://euw1.api.riotgames.com/lol/match/v4/matchlists/by-account/AFgU6nAb0MBYMUs9eeOmh_rauERBgTb0DSTJQveFYLeZ02o?queue=420&season=13&endTime=1583290947037&beginTime=1582876147000&endIndex=100&beginIndex=0&api_key=RGAPI-f5fdf9de-705b-4e99-92fd-0e731b2339bd
#https://euw1.api.riotgames.com/lol/match/v4/matchlists/by-account/+id+?queue=420&season=13&endTime=+x+&beginTime=+x+&endIndex=100&beginIndex=0&api_key=+x+
#summonerid
#https://euw1.api.riotgames.com/lol/summoner/v4/summoners/nkL9hjOo6j8o2sZF5YdcyM08SNbcQXny871OHCyV8PLr5W0?api_key=RGAPI-f5fdf9de-705b-4e99-92fd-0e731b2339bd
#https://euw1.api.riotgames.com/lol/summoner/v4/summoners/+id+?api_key=+apikey+
#matches
#https://euw1.api.riotgames.com/lol/match/v4/matches/4435844962?api_key=RGAPI-f5fdf9de-705b-4e99-92fd-0e731b2339bd
#https://euw1.api.riotgames.com/lol/match/v4/matches/+matchid+?api_key=+apikey+
#output
#[4429683428[0], 68[1], 10.4[2], ['Lukezy[3][0]', None[3][1], 'lukezy'[3][2], 'Lukezy'[3][3], '1wxnkD9N7Qg2r5eIX6nEpgoEDIJcCucC70k-Db7F8pIKoLA'[3][4]], 0[4]

def main():
    apikey = input('APIKey: ')
    #apikey = 'RGAPI-371cb3dd-cab0-4492-901a-879e6df78489'
    #teamname = 'SK Gaming'
    teamname = input('Team Name: ')
    wboutput = Workbook()
    dest_filename = teamname + '.xlsx'
    ws1 = wboutput.active
    ws1.title = "Players"
    #wb = load_workbook(filename = 'SKGamingaccs.xlsx')
    wb = load_workbook(filename = 'database.xlsx')
    sheet_ranges = wb['Sheet1']
    patchlist = [[10.1, 1578441600000], [10.2, 1579737600000], [10.3, 1580860800000], [10.4, 1582156800000], [10.5, 1583280000000], [10.6, 1584489600000], [10.7, 1585699200000], [10.8, 1586908800000], [10.9, 1588118400000], [10.1, 1589328000000], [10.11, 1590624000000], [10.12, 1591747200000], [10.13, 1592956800000], [10.14, 1594166400000], [10.15, 1595376000000], [10.16, 1596585600000], [10.17, 1597795200000], [10.18, 1599004800000], [10.19, 1600214400000], [10.2, 1601424000000], [10.21, 1602633600000], [10.22, 1603843200000], [10.23, 1605052800000], [10.24, 1606176000000], [10.25, 1607472000000]]
    players = playernamegrab(teamname, sheet_ranges)
    print(players)
    startdate = arrow.utcnow()
    startdate = epoch(startdate)
    print(startdate)
    patcharray = patchdates(startdate, patchlist)
    print(patcharray)
    results = []
    for player in players:
        if players[players.index(player)] =='END':
            break
        for patchdate in patcharray:
            if patcharray[patcharray.index(patchdate)+1] == 'END':
                break
            results += matchhistory(apikey, idconverter(players[players.index(player)][4], apikey), patcharray[patcharray.index(patchdate)], patcharray[patcharray.index(patchdate)+1], patchlist, player)
    print(results)
    outputprint(teamname, results, ws1, dest_filename, wboutput, apikey)
    print('Players listed')

def outputprint(teamname, results, ws1, dest_filename, wboutput, apikey):
    rowdata = 1
    URL = "http://ddragon.leagueoflegends.com/cdn/10.5.1/data/en_US/champion.json"
    response = requests.get(URL)
    response = response.json()
    champions = response['data']
    ws1.cell(column=1, row=1, value="Team")
    ws1.cell(column=2, row=1, value="Player")
    ws1.cell(column=3, row=1, value="Role")
    ws1.cell(column=4, row=1, value="Previous ign")
    ws1.cell(column=5, row=1, value="Current ign")
    ws1.cell(column=6, row=1, value="Patch")
    ws1.cell(column=7, row=1, value="Champion id")
    ws1.cell(column=8, row=1, value="Result")
    ws1.cell(column=9, row=1, value="Match id")
    for x in results:
        print(results[results.index(x)][3][0])
        rowdata += 1
        ws1.cell(column=1, row=rowdata, value=teamname)
        ws1.cell(column=2, row=rowdata, value=results[results.index(x)][3][0])
        ws1.cell(column=3, row=rowdata, value=results[results.index(x)][3][1])
        ws1.cell(column=4, row=rowdata, value=results[results.index(x)][3][2])
        ws1.cell(column=5, row=rowdata, value=results[results.index(x)][3][3])
        ws1.cell(column=6, row=rowdata, value=results[results.index(x)][2])
        for a in champions:
            if int(champions[a]['key']) == int(results[results.index(x)][1]):
                print("found")
                ws1.cell(column=7, row=rowdata, value=champions[a]['id'])
        try:
            ws1.cell(column=8, row=rowdata, value=results[results.index(x)][4])
        except:
            try:
                ws1.cell(column=8, row=rowdata, value=matchfix(results[results.index(x)][0], results[results.index(x)][1], apikey))
            except:
                pass
        ws1.cell(column=9, row=rowdata, value=results[results.index(x)][0])
    wboutput.save(filename = dest_filename)

def playernamegrab(teamname, sheet_ranges):
    xx = 2
    players = []
    while str(sheet_ranges['A'+str(xx)].value) != 'None':
        if sheet_ranges['A'+str(xx)].value == teamname:
            players.append([sheet_ranges['B'+str(xx)].value, sheet_ranges['C'+str(xx)].value, sheet_ranges['D'+str(xx)].value, sheet_ranges['E'+str(xx)].value, sheet_ranges['F'+str(xx)].value])
        xx += 1
    players.append('END')
    print('playersgrabbed')
    return players

def epoch(now):
    s = now.timestamp
    ms = int(now.format("SSS"))
    string = (s * 1000 + ms)
    return string

def reverseepoch(time):
    time = str(time)
    time = time[:10]
    print(time)
    return time

def idconverter(playerid, apikey):
    URL = "https://euw1.api.riotgames.com/lol/summoner/v4/summoners/" + playerid + "?api_key=" + apikey
    response = requests.get(URL)
    response = response.json()
    try:
        if response['status']['status_code'] == 429 or response['status']['status_code'] == 504 or response['status']['status_code'] == 503:
            print('timeout (60s)')
            print(response)
            if response['status']['status_code'] == 429:
                print('429')
                time.sleep(60)
            if response['status']['status_code'] == 500:
                print('500')
                time.sleep(120)
            if response['status']['status_code'] == 503:
                print('503')
                time.sleep(120)
            if response['status']['status_code'] == 504:
                print('504')
                time.sleep(120)
            URL = "https://euw1.api.riotgames.com/lol/summoner/v4/summoners/" + playerid + "?api_key=" + apikey
            response = requests.get(URL)
            response = response.json()
            if response['status']['status_code'] == 429 or response['status']['status_code'] == 504 or response['status']['status_code'] == 503:
                print('timeout (60s)')
                time.sleep(60)
                URL = "https://euw1.api.riotgames.com/lol/summoner/v4/summoners/" + playerid + "?api_key=" + apikey
                response = requests.get(URL)
                response = response.json()
            elif response['status']['status_code'] == 404:
                print(response['status']['status_code'])
            elif response['status']['status_code'] != 200:
                print(response['status']['status_code'])
        elif response['status']['status_code'] == 404:
            print(response['status']['status_code'])
        elif response['status']['status_code'] == 401:
            print(response['status']['status_code'])
            print('Player id error, apikey doesnt match')
            exit()
        elif response['status']['status_code'] == 403:
            print(response['status']['status_code'])
            print('Request forbidden, check apikey')
            exit()
        elif response['status']['status_code'] != 200:
            print(response['status']['status_code'])
    except:
        pass
    print(response)
    playerid = response['accountId']
    playerid = str(playerid)
    return playerid

def matchgrabber(apikey, players, startdate, enddate, patchlist):
    matcharray = []
    for player in players:
        matcharray.append(matchhistory(apikey, idconverter(players[players.index(player)][4], apikey), startdate, enddate, patchlist, player))
    return matcharray

def patchdates(currentdate, patchlist):
    datearray = []
    now = arrow.utcnow()
    now = now.shift(weeks=-1)
    datearray.append(currentdate)
    for x in patchlist:
        if patchlist[patchlist.index(x)][1] < currentdate and patchlist[patchlist.index(x)+1][1] > currentdate:
            if epoch(now) < patchlist[patchlist.index(x)][1]:
                datearray.append(patchlist[patchlist.index(x)][1])
            else:
                datearray.append(epoch(now))
                datearray.append(patchlist[patchlist.index(x)][1])
            now = arrow.Arrow.fromtimestamp(reverseepoch(patchlist[patchlist.index(x)][1]))
            print(now)
            datearray.append(epoch(now.shift(weeks=-1)))
            datearray.append(patchlist[patchlist.index(x)-1][1])
            datearray.append('END')
    return datearray

def matchhistory(apikey, playerid, startdate, enddate, patchlist, player):
    matcharray = []
    matchids = []
    URL = "https://euw1.api.riotgames.com/lol/match/v4/matchlists/by-account/" + str(playerid) + "?queue=420&season=13&endTime=" + str(startdate) + "&beginTime=" + str(enddate) + "&endIndex=100&beginIndex=0&api_key=" + apikey
    response = requests.get(URL)
    response = response.json()
    print(response)
    try:
        if response['status']['status_code'] == 429 or response['status']['status_code'] == 504 or response['status']['status_code'] == 503:
            print('timeout (60s)')
            print(response)
            if response['status']['status_code'] == 429:
                print('429')
                time.sleep(60)
            if response['status']['status_code'] == 500:
                print('500')
                time.sleep(120)
            if response['status']['status_code'] == 503:
                print('503')
                time.sleep(120)
            if response['status']['status_code'] == 504:
                print('504')
                time.sleep(120)
            URL = "https://euw1.api.riotgames.com/lol/match/v4/matchlists/by-account/" + str(playerid) + "?queue=420&season=13&endTime=" + str(startdate) + "&beginTime=" + str(enddate) + "&endIndex=100&beginIndex=0&api_key=" + apikey
            response = requests.get(URL)
            response = response.json()
            if response['status']['status_code'] == 429 or response['status']['status_code'] == 504 or response['status']['status_code'] == 503:
                print('timeout (60s)')
                time.sleep(60)
                URL = "https://euw1.api.riotgames.com/lol/match/v4/matchlists/by-account/" + str(playerid) + "?queue=420&season=13&endTime=" + str(startdate) + "&beginTime=" + str(enddate) + "&endIndex=100&beginIndex=0&api_key=" + apikey
                response = requests.get(URL)
                response = response.json()
            elif response['status']['status_code'] == 404:
                print(response['status']['status_code'])
            elif response['status']['status_code'] == 400:
                print('400')
                URL = "https://euw1.api.riotgames.com/lol/match/v4/matchlists/by-account/" + str(idconverter(playerid)) + "?queue=420&season=13&endTime=" + str(startdate) + "&beginTime=" + str(enddate) + "&endIndex=100&beginIndex=0&api_key=" + apikey
                response = requests.get(URL)
                response = response.json()
            elif response['status']['status_code'] != 200:
                print(response['status']['status_code'])
        elif response['status']['status_code'] == 400:
            print('400')
            URL = "https://euw1.api.riotgames.com/lol/match/v4/matchlists/by-account/" + str(idconverter(playerid)) + "?queue=420&season=13&endTime=" + str(startdate) + "&beginTime=" + str(enddate) + "&endIndex=100&beginIndex=0&api_key=" + apikey
            response = requests.get(URL)
            response = response.json()
        elif response['status']['status_code'] == 404:
            print(response['status']['status_code'])
            return matcharray
        elif response['status']['status_code'] == 401:
            print(response['status']['status_code'])
            print('Player id error, apikey doesnt match')
            exit()
        elif response['status']['status_code'] == 403:
            print(response['status']['status_code'])
            print('Request forbidden, check apikey')
            exit()
        elif response['status']['status_code'] != 200:
            print(response['status']['status_code'])
    except:
            pass
    try:
        matches = response['matches']
    except:
        print(response)
        print(URL)
        print("ERROR2")
        return matcharray
    for match in matches:
        matchids.append([match['gameId'], match['champion'], patchfilter(match['timestamp'], patchlist), player])
    matcharray = matchget(matchids, apikey)
    print('Matchesgrabbed')
    return matcharray

def patchfilter(timestamp, patchlist):
    for x in patchlist:
        if patchlist[patchlist.index(x)][1] < timestamp and patchlist[patchlist.index(x)+1][1] > timestamp:
            patch = patchlist[patchlist.index(x)][0]
    return patch

def matchget(matchids, apikey):
    for x in matchids:
        URL = "https://euw1.api.riotgames.com/lol/match/v4/matches/" + str(matchids[matchids.index(x)][0]) + "?api_key=" + apikey
        response2 = requests.get(URL)
        response2 = response2.json()
        try:
            if response2['status']['status_code'] == 429 or response2['status']['status_code'] == 504 or response2['status']['status_code'] == 503:
                print('timeout (60s)')
                print(response2)
                if response2['status']['status_code'] == 429:
                    print('429')
                if response2['status']['status_code'] == 503:
                    print('503')
                    time.sleep(60)
                if response2['status']['status_code'] == 504:
                    print('504')
                    time.sleep(60)
                time.sleep(60)
                URL = "https://euw1.api.riotgames.com/lol/match/v4/matches/" + str(matchids[matchids.index(x)][0]) + "?api_key=" + apikey
                response2 = requests.get(URL)
                response2 = response2.json()
                try:
                    if response2['status']['status_code'] == 429 or response2['status']['status_code'] == 504 or response2['status']['status_code'] == 503:
                        print('timeout (60s)')
                        time.sleep(60)
                        URL = "https://euw1.api.riotgames.com/lol/match/v4/matches/" + str(matchids[matchids.index(x)][0]) + "?api_key=" + apikey
                        response2 = requests.get(URL)
                        response2 = response.json()
                    elif response2['status']['status_code'] == 404:
                        print(response2['status']['status_code'])
                        continue
                    elif response2['status']['status_code'] != 200:
                        print(response2['status']['status_code'])
                except:
                    pass
            elif response2['status']['status_code'] == 404:
                print(response2['status']['status_code'])
                continue
            elif response2['status']['status_code'] == 401:
                print(response['status']['status_code'])
                print('Player id error, apikey doesnt match')
                exit()
            elif response2['status']['status_code'] == 403:
                print(response['status']['status_code'])
                print('Request forbidden, check apikey')
                exit()
            elif response2['status']['status_code'] != 200:
                print(response2['status']['status_code'])
        except:
            pass
        try:
            participants = response2['participants']
            for z in participants:
                if z['championId'] == matchids[matchids.index(x)][1]:
                    if z['stats']['win']:
                        matchids[matchids.index(x)] += [1]
                    else:
                        matchids[matchids.index(x)] += [0]
        except:
            print('Error')
            print(response2)
            pass
    return matchids


def matchfix(matchid, championid, apikey):
    URL = "https://euw1.api.riotgames.com/lol/match/v4/matches/" + str(matchid) + "?api_key=" + apikey
    response2 = requests.get(URL)
    response2 = response2.json()
    try:
        if response['status']['status_code'] == 429 or response['status']['status_code'] == 504 or response['status']['status_code'] == 503:
            print('timeout (60s)')
            print(response)
            if response['status']['status_code'] == 429:
                print('429')
                time.sleep(60)
            if response['status']['status_code'] == 500:
                print('500')
                time.sleep(120)
            if response['status']['status_code'] == 503:
                print('503')
                time.sleep(120)
            if response['status']['status_code'] == 504:
                print('504')
                time.sleep(120)
            URL = "https://euw1.api.riotgames.com/lol/match/v4/matches/" + str(matchid) + "?api_key=" + apikey
            response = requests.get(URL)
            response = response.json()
            if response['status']['status_code'] == 429 or response['status']['status_code'] == 504 or response['status']['status_code'] == 503:
                print('timeout (60s)')
                time.sleep(60)
                URL = "https://euw1.api.riotgames.com/lol/match/v4/matches/" + str(matchid) + "?api_key=" + apikey
                response = requests.get(URL)
                response = response.json()
            elif response['status']['status_code'] == 404:
                print(response['status']['status_code'])
            elif response['status']['status_code'] != 200:
                print(response['status']['status_code'])
        elif response2['status']['status_code'] == 404:
            print(response2['status']['status_code'])
        elif response2['status']['status_code'] != 200:
            print(response2['status']['status_code'])
        participants = response2['participants']
        for z in participants:
            if z['championId'] == championid:
                if z['stats']['win']:
                    return 1
                else:
                    return 0
    except:
        print('Error')
        print(response2)
        return 0

if __name__ == '__main__':
    main()

